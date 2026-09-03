-- 苏E好学 Supabase 导入脚本
-- Run this after the schema is created. It uses the local Excel export.

import json
import os
import re
import psycopg2
from psycopg2.extras import execute_values
from pathlib import Path
import zipfile
import io
from openpyxl import load_workbook

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
BATCH_ID = "2026秋季"  # 固定批次ID

if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
    print("请设置 SUPABASE_URL 和 SUPABASE_SERVICE_ROLE_KEY 环境变量")
    exit(1)

conn = psycopg2.connect(
    f"postgresql://{SUPABASE_URL.split('//')[1]}/{SUPABASE_SERVICE_ROLE_KEY}"
)
cur = conn.cursor()

def upsert(table, data, conflict_keys, returning=None):
    cols = list(data[0].keys())
    placeholders = ", ".join([f"%({k})s" for k in cols])
    conflict = " ON CONFLICT (" + ", ".join(conflict_keys) + ") DO UPDATE SET " + ", ".join([f"{k}=EXCLUDED.{k}" for k in cols if k not in conflict_keys])
    query = f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({placeholders}) {conflict} RETURNING {returning or 'id'}"
    execute_values(cur, query, data)
    conn.commit()

def parse_roster(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {
            "班级名称": row[0],
            "年级": row[1],
            "学科": row[2],
            "学期": row[3],
            "校区": row[4],
            "老师": row[5],
            "开课日期": row[6],
            "结束日期": row[7],
            "讲次时间": row[8],
            "已报/预招": row[9],
            "学生ID": row[10],
            "学生姓名": row[11],
            "英文名": row[12],
            "联系电话": row[13],
            "性别": row[14],
            "学生年级": row[15],
            "状态": row[16],
            "课程费应收": row[17],
            "书本费应收": row[18],
            "欠费金额": row[19]
        }
        rows.append(row_dict)
    return rows

def parse_zip(zip_path):
    with zipfile.ZipFile(zip_path) as z:
        for info in z.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            data = io.BytesIO(z.read(info))
            wb = load_workbook(data, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                row_dict = {
                    "班级名称": row[0],
                    "年级": row[1],
                    "学科": row[2],
                    "学期": row[3],
                    "校区": row[4],
                    "老师": row[5],
                    "开课日期": row[6],
                    "结束日期": row[7],
                    "讲次时间": row[8],
                    "已报/预招": row[9],
                    "学生ID": row[10],
                    "学生姓名": row[11],
                    "英文名": row[12],
                    "联系电话": row[13],
                    "性别": row[14],
                    "学生年级": row[15],
                    "状态": row[16],
                    "课程费应收": row[17],
                    "书本费应收": row[18],
                    "欠费金额": row[19]
                }
                yield row_dict, info.filename

def normalize_class_name(cls):
    return re.sub(r'[-]', '', cls).strip()

# 1. 导入批次
raw_data = parse_roster("C:/Users/QQQ/Downloads/班级学生导出文件20260903201614/班级学生汇总.xlsx")
data = [
    {
        "班级名称": r["班级名称"],
        "年级": r["年级"],
        "学科": r["学科"],
        "学期": r["学期"],
        "校区": r["校区"],
        "老师": r["老师"],
        "开课日期": r["开课日期"],
        "结束日期": r["结束日期"],
        "讲次时间": r["讲次时间"],
        "已报/预招": r["已报/预招"],
        "学生ID": r["学生ID"],
        "学生姓名": r["学生姓名"],
        "英文名": r["英文名"],
        "联系电话": r["联系电话"],
        "性别": r["性别"],
        "学生年级": r["学生年级"],
        "状态": r["状态"],
        "课程费应收": r["课程费应收"],
        "书本费应收": r["书本费应收"],
        "欠费金额": r["欠费金额"]
    }
    for r in raw_data
]

upsert("raw_roster_rows", data, ["batch_id", "source_file", "row_index"], "id")

# 2. 班级
classes = {}
for r in raw_data:
    cls = r["班级名称"]
    if cls not in classes:
        classes[cls] = {
            "class_name": cls,
            "normalized_class_name": normalize_class_name(cls),
            "term": "2026秋",
            "grade": r["年级"],
            "subject": r["学科"],
            "campus": r["校区"],
            "teacher": r["老师"],
            "start_date": r["开课日期"],
            "end_date": r["结束日期"],
            "raw": {"原始": r}
        }

upsert("classes", list(classes.values()), ["class_name"])

# 3. 学生（先按 sourceStudentId 去重）
students_map = {}
for r in raw_data:
    sid = r["学生ID"]
    if sid not in students_map:
        students_map[sid] = {
            "id": sid,
            "name": r["学生姓名"],
            "phone": r["联系电话"],
            "gender": r["性别"],
            "grade": r["学生年级"],
            "english_name": r["英文名"],
            "source_student_id": sid,
            "family_id": None,  # 后续家庭表有后补
            "raw": {"原始": r}
        }

upsert("students", list(students_map.values()), ["id"])

# 4. 家庭（先按 phone 去重）
families_map = {}
for r in raw_data:
    phone = r["联系电话"]
    if phone and phone not in families_map:
        families_map[phone] = {
            "phone": phone,
            "source_name": r["学生姓名"],
            "raw": {"原始": r}
        }

upsert("families", list(families_map.values()), ["phone"])

# 5. 报名（用最新 Excel 覆盖）
enrollments = []
for r in raw_data:
    enrollments.append({
        "student_id": r["学生ID"],
        "class_name": r["班级名称"],
        "normalized_class_name": normalize_class_name(r["班级名称"]),
        "term": "2026秋",
        "grade": r["年级"],
        "subject": r["学科"],
        "campus": r["校区"],
        "teacher": r["老师"],
        "start_date": r["开课日期"],
        "end_date": r["结束日期"],
        "source_status": r["状态"],
        "amount_due": r["课程费应收"],
        "book_fee": r["书本费应收"],
        "arrears": r["欠费金额"],
        "raw": {"原始": r}
    })

upsert("enrollments", enrollments, ["student_id", "class_name", "term"])

print("导入完成")
