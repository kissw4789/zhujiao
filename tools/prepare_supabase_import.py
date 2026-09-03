# -*- coding: utf-8 -*-
"""Prepare Supabase import payloads from the latest roster export.

This script is local-only. It writes JSON payloads and a dry-run report to the
archive directory. It does not connect to Supabase and does not require secrets.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXPORT_DIR = Path(r"C:/Users/QQQ/Downloads/班级学生导出文件20260903201614")
LOCAL_BACKEND_DATA_DIR = Path(r"C:/助教系统/助教网页端/苏E好学/施工区/后台网站/data")
DEPLOY_DATA_DIR = ROOT / "data"
DEFAULT_ARCHIVE_DIR = ROOT.parent / "_archive" / "2026-09-03_supabase_migration"

def pick_data_dir() -> Path:
    candidates = [LOCAL_BACKEND_DATA_DIR, DEPLOY_DATA_DIR]
    scored: list[tuple[int, Path]] = []
    for p in candidates:
        if not p.exists():
            continue
        score = 0
        for name in ["students.json", "families.json", "enrollments.json", "orders.json", "schedule.json", "state.json", "outlines.json"]:
            fp = p / name
            if fp.exists():
                score += fp.stat().st_size
        scored.append((score, p))
    if scored:
        scored.sort(key=lambda x: x[0], reverse=True)
        return scored[0][1]
    return DEPLOY_DATA_DIR

DEFAULT_DATA_DIR = pick_data_dir()

ROSTER_FILE = "班级学生汇总.xlsx"
DETAIL_ZIP = "班级学生明细.zip"

HEADERS = [
    "班级名称", "年级", "学科", "学期", "校区", "老师", "开课日期", "结束日期", "讲次时间", "已报/预招",
    "学生ID", "学生姓名", "英文名", "联系电话", "性别", "学生年级", "状态", "课程费应收", "书本费应收", "欠费金额",
]

TEACHER_ALIASES = {
    "飞飞": "王易飞", "温温": "温佳炜", "小明": "小明老师", "小明老师": "小明老师", "小天": "陈世崇",
    "小树": "束亚成", "金金": "刘金鑫", "晓晓": "张梦晓", "章章": "章雪萍", "俞老师": "俞锐钦",
}


def s(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def dstr(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = s(value)
    if not text:
        return ""
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return text


def money(value: Any) -> float:
    text = s(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def term_of(start_date: str, term_name: str = "") -> str:
    if re.match(r"^\d{4}[春暑秋寒]$", s(term_name)):
        return s(term_name)
    if "秋" in s(term_name):
        y = start_date[:4] if re.match(r"^\d{4}", start_date) else "2026"
        return f"{y}秋"
    if "暑" in s(term_name):
        y = start_date[:4] if re.match(r"^\d{4}", start_date) else "2026"
        return f"{y}暑"
    if "春" in s(term_name):
        y = start_date[:4] if re.match(r"^\d{4}", start_date) else "2026"
        return f"{y}春"
    if start_date:
        y = int(start_date[:4])
        m = int(start_date[5:7])
        if 3 <= m <= 5:
            return f"{y}春"
        if 6 <= m <= 8:
            return f"{y}暑"
        if 9 <= m <= 11:
            return f"{y}秋"
        return f"{y}寒"
    return "2026秋"


def normalize_teacher(value: str) -> str:
    text = s(value).replace("老师", "")
    return TEACHER_ALIASES.get(text) or TEACHER_ALIASES.get(text + "老师") or s(value)


def normalize_subject(value: str, class_name: str = "") -> str:
    text = s(value) + s(class_name)
    if "物理" in text:
        return "物理"
    if any(k in text for k in ["数学", "奥数", "中考", "自招", "创新", "尖子", "小明"]):
        return "数学"
    return s(value) or "数学"


def class_type(class_name: str) -> str:
    text = s(class_name)
    if "小明班" in text:
        return "小明班"
    if "自招" in text:
        return "自招"
    if "中考" in text:
        return "中考"
    if "创新" in text:
        return "创新"
    if "尖子" in text:
        return "尖子"
    if "奥综" in text or "奥数" in text:
        return "奥数"
    return "其他"


def normalized_class_name(class_name: str) -> str:
    text = s(class_name)
    text = text.replace("自招-A", "自招A").replace("自招-B", "自招B")
    text = re.sub(r"\s+", "", text)
    return text


def stable_hash(*parts: Any, length: int = 12) -> str:
    raw = "|".join(s(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:length]


def read_xlsx_rows(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    header = [s(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, str]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(s(v) for v in row):
            continue
        item = {header[i]: s(row[i]) if i < len(row) else "" for i in range(len(header))}
        for key in ["开课日期", "结束日期"]:
            col = header.index(key) if key in header else -1
            if col >= 0 and col < len(row):
                item[key] = dstr(row[col])
        item["_row_index"] = str(idx)
        rows.append(item)
    return rows


def source_class_map_from_zip(path: Path) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not path.exists():
        return mapping
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            if info.is_dir() or not info.filename.lower().endswith(".xlsx"):
                continue
            base = Path(info.filename).stem
            m = re.search(r"-(\d+)$", base)
            source_class_id = m.group(1) if m else stable_hash(base)
            with zf.open(info) as fp:
                wb = load_workbook(fp, data_only=True, read_only=False)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    class_name = s(row[0]) if row else ""
                    if class_name:
                        mapping.setdefault(normalized_class_name(class_name), source_class_id)
    return mapping


def load_json(path: Path, default: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def cn_student(row: Dict[str, str], existing: Dict[str, Any] | None = None) -> Dict[str, Any]:
    sid = s(row.get("学生ID"))
    name = s(row.get("学生姓名"))
    grade = s(row.get("学生年级"))
    if not grade or grade == "无年级":
        grade = s(row.get("年级"))
    existing = existing or {}
    return {
        "id": existing.get("id") or sid,
        "family_id": existing.get("familyId") or existing.get("family_id") or f"F-{sid}",
        "source_student_id": existing.get("sourceStudentId") or existing.get("source_student_id") or sid,
        "name": name,
        "phone": s(row.get("联系电话")) or existing.get("电话") or existing.get("phone") or "",
        "gender": s(row.get("性别")) or existing.get("性别") or existing.get("gender") or "",
        "grade": grade,
        "english_name": s(row.get("英文名")) or existing.get("英文名") or existing.get("english_name") or "",
        "tags": existing.get("标签") or existing.get("tags") or [],
        "intent": existing.get("意向") or existing.get("intent") or "",
        "note": existing.get("备注") or existing.get("note") or "",
        "family_order": int(existing.get("家庭排序") or existing.get("family_order") or 1),
        "first_date": existing.get("首次") or existing.get("first_date") or s(row.get("开课日期")),
        "recent_date": existing.get("最近") or existing.get("recent_date") or s(row.get("开课日期")),
        "enrollment_count": int(existing.get("次数") or existing.get("enrollment_count") or 0),
        "source_name": existing.get("来源姓名") or existing.get("source_name") or name,
        "assignment_confirmed": bool(existing.get("分配确认") or existing.get("assignment_confirmed") or False),
        "is_manual": bool(existing.get("手工") or existing.get("is_manual") or False),
        "raw": {"latest_export": row, "legacy": existing},
    }


def legacy_student_maps(data_dir: Path) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    students = load_json(data_dir / "students.json", [])
    families = load_json(data_dir / "families.json", [])
    by_id = {s(item.get("id")): item for item in students if item.get("id")}
    by_source = defaultdict(list)
    for item in students:
        key = s(item.get("sourceStudentId")) or s(item.get("id"))
        if key:
            by_source[key].append(item)
    fam_by_id = {s(item.get("familyId")): item for item in families if item.get("familyId")}
    return by_id, by_source, fam_by_id


def build_payloads(export_dir: Path, data_dir: Path) -> Dict[str, Any]:
    rows = read_xlsx_rows(export_dir / ROSTER_FILE)
    class_sources = source_class_map_from_zip(export_dir / DETAIL_ZIP)
    by_student_id, by_source, fam_by_id = legacy_student_maps(data_dir)
    legacy_families = load_json(data_dir / "families.json", [])
    legacy_enrollments = load_json(data_dir / "enrollments.json", [])
    legacy_orders = load_json(data_dir / "orders.json", [])
    legacy_schedule = load_json(data_dir / "schedule.json", [])
    legacy_state = load_json(data_dir / "state.json", {})
    outlines = load_json(data_dir / "outlines.json", {})

    batch_id = "20260903201614"
    batch_summary = {
        "口径": "以最新班级学生汇总.xlsx 为 2026秋主口径，保留旧状态和人工记录",
        "source_export_dir": str(export_dir),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }

    families: Dict[str, Dict[str, Any]] = {}
    for f in legacy_families:
        fid = s(f.get("familyId"))
        if not fid:
            continue
        families[fid] = {
            "family_id": fid,
            "phone": s(f.get("phone")),
            "source_student_id": s(f.get("sourceStudentId")),
            "source_name": s(f.get("sourceName")),
            "campus": s(f.get("campus")),
            "source_status": s(f.get("sourceStatus")),
            "needs_review": bool(f.get("needsReview") or False),
            "pending_enrollments": int(f.get("pendingEnrollments") or 0),
            "cluster_summary": f.get("clusterSummary") or {},
            "last_sync_at": s(f.get("lastSyncAt")),
            "migration_version": s(f.get("migrationVersion")),
            "raw": f,
        }

    students: Dict[str, Dict[str, Any]] = {}
    for sid, legacy in by_student_id.items():
        if not sid:
            continue
        students[sid] = {
            "id": sid,
            "family_id": s(legacy.get("familyId")) or f"F-{sid}",
            "source_student_id": s(legacy.get("sourceStudentId")) or sid,
            "name": s(legacy.get("姓名")) or sid,
            "phone": s(legacy.get("电话")),
            "gender": s(legacy.get("性别")),
            "grade": s(legacy.get("年级")),
            "english_name": s(legacy.get("英文名")),
            "tags": legacy.get("标签") or [],
            "intent": s(legacy.get("意向")),
            "note": s(legacy.get("备注")),
            "family_order": int(legacy.get("家庭排序") or 1),
            "first_date": s(legacy.get("首次")),
            "recent_date": s(legacy.get("最近")),
            "enrollment_count": int(legacy.get("次数") or 0),
            "source_name": s(legacy.get("来源姓名")),
            "assignment_confirmed": bool(legacy.get("分配确认") or False),
            "is_manual": bool(legacy.get("手工") or False),
            "raw": legacy,
        }

    classes: Dict[str, Dict[str, Any]] = {}
    enrollments: Dict[str, Dict[str, Any]] = {}
    raw_roster_rows: List[Dict[str, Any]] = []
    latest_enrollment_keys = set()

    for pos, row in enumerate(rows, start=1):
        sid = s(row.get("学生ID"))
        class_name = s(row.get("班级名称"))
        nclass = normalized_class_name(class_name)
        source_class_id = class_sources.get(nclass)
        start = s(row.get("开课日期"))
        end = s(row.get("结束日期"))
        term = term_of(start, s(row.get("学期")))
        class_id = source_class_id or stable_hash(nclass, term)
        cid = f"CLS-{class_id}"
        grade = s(row.get("年级"))
        subject = normalize_subject(s(row.get("学科")), class_name)
        lecture_times = s(row.get("讲次时间"))
        times = [x.strip() for x in lecture_times.split(",") if x.strip()]
        time_range = times[0] if times else ""
        teacher = normalize_teacher(s(row.get("老师")))

        source_students = by_source.get(sid) or []
        legacy = source_students[0] if len(source_students) == 1 else by_student_id.get(sid, {})
        student_id = s(legacy.get("id")) or sid
        student = cn_student(row, legacy)
        student["id"] = student_id
        if student["family_id"] not in families:
            families[student["family_id"]] = {
                "family_id": student["family_id"],
                "phone": student["phone"],
                "source_student_id": sid,
                "source_name": s(row.get("学生姓名")),
                "campus": s(row.get("校区")),
                "source_status": "latest_export",
                "needs_review": "/" in s(row.get("学生姓名")) or "／" in s(row.get("学生姓名")),
                "pending_enrollments": 0,
                "cluster_summary": {},
                "last_sync_at": "",
                "migration_version": "supabase-20260903",
                "raw": {"latest_export": row},
            }
        students[student_id] = {**students.get(student_id, {}), **student}

        classes[cid] = {
            "id": cid,
            "source_class_id": source_class_id,
            "class_name": class_name,
            "normalized_class_name": nclass,
            "term": term,
            "term_name": s(row.get("学期")) or term,
            "grade": grade,
            "subject": subject,
            "campus": s(row.get("校区")),
            "teacher": teacher,
            "start_date": start,
            "end_date": end,
            "weekday": "",
            "time_range": time_range,
            "lecture_times": lecture_times,
            "capacity_text": s(row.get("已报/预招")),
            "class_type": class_type(class_name),
            "raw": row,
            "last_import_batch": batch_id,
            "active_in_latest": True,
        }

        eid = f"E-{student_id}-{class_id}"
        latest_enrollment_keys.add(eid)
        enrollments[eid] = {
            "eid": eid,
            "student_id": student_id,
            "source_student_id": sid,
            "family_id": students[student_id]["family_id"],
            "class_id": cid,
            "student_name": s(row.get("学生姓名")),
            "phone": s(row.get("联系电话")),
            "class_name": class_name,
            "class_display_name": class_name,
            "normalized_class_name": nclass,
            "grade": grade,
            "subject": subject,
            "term": term,
            "term_name": s(row.get("学期")) or term,
            "campus": s(row.get("校区")),
            "teacher": teacher,
            "time_range": time_range,
            "lecture_times": lecture_times,
            "start_date": start,
            "end_date": end,
            "source_status": s(row.get("状态")),
            "assignment_status": "已确认" if student_id else "待确认",
            "display_status": s(row.get("状态")),
            "capacity_text": s(row.get("已报/预招")),
            "amount_due": money(row.get("课程费应收")),
            "book_fee": money(row.get("书本费应收")),
            "arrears": money(row.get("欠费金额")),
            "fee_text": s(row.get("课程费应收")),
            "is_void": False,
            "is_manual": False,
            "active_in_latest": s(row.get("状态")) != "历史在班学生",
            "last_import_batch": batch_id,
            "raw": row,
        }
        raw_roster_rows.append({
            "batch_id": batch_id,
            "source_file": ROSTER_FILE,
            "row_index": pos,
            "source_class_id": source_class_id,
            "source_student_id": sid,
            "class_name": class_name,
            "normalized_class_name": nclass,
            "payload": row,
        })

    # Preserve only non-latest-term or manual legacy enrollments. 2026秋 ordinary roster rows are governed by the latest export.
    for legacy in legacy_enrollments:
        legacy_term = s(legacy.get("期")) or term_of(s(legacy.get("开课")), s(legacy.get("学期")))
        keep_legacy = legacy_term != "2026秋" or bool(legacy.get("手工")) or bool(legacy.get("作废"))
        if not keep_legacy:
            continue
        legacy_eid = s(legacy.get("eid"))
        if not legacy_eid:
            legacy_eid = f"LEG-{stable_hash(legacy.get('id') or legacy.get('studentId'), legacy.get('班级'), legacy.get('开课'))}"
        if legacy_eid in enrollments:
            continue
        sid = s(legacy.get("childId") or legacy.get("id") or legacy.get("studentId"))
        class_name = s(legacy.get("班级") or legacy.get("班级名称"))
        nclass = normalized_class_name(class_name)
        class_id = f"CLS-{stable_hash(nclass, legacy.get('期') or legacy.get('开课'))}"
        if sid and sid not in students and sid in by_student_id:
            old = by_student_id[sid]
            students[sid] = {
                "id": sid,
                "family_id": s(old.get("familyId")) or f"F-{sid}",
                "source_student_id": s(old.get("sourceStudentId")) or sid,
                "name": s(old.get("姓名")) or sid,
                "phone": s(old.get("电话")),
                "gender": s(old.get("性别")),
                "grade": s(old.get("年级")),
                "english_name": s(old.get("英文名")),
                "tags": old.get("标签") or [],
                "intent": s(old.get("意向")),
                "note": s(old.get("备注")),
                "family_order": int(old.get("家庭排序") or 1),
                "first_date": s(old.get("首次")),
                "recent_date": s(old.get("最近")),
                "enrollment_count": int(old.get("次数") or 0),
                "source_name": s(old.get("来源姓名")),
                "assignment_confirmed": bool(old.get("分配确认") or False),
                "is_manual": bool(old.get("手工") or False),
                "raw": old,
            }
        if class_name and class_id not in classes:
            classes[class_id] = {
                "id": class_id,
                "source_class_id": None,
                "class_name": class_name,
                "normalized_class_name": nclass,
                "term": s(legacy.get("期")) or term_of(s(legacy.get("开课")), s(legacy.get("学期"))),
                "term_name": s(legacy.get("学期")),
                "grade": s(legacy.get("年级")),
                "subject": normalize_subject(s(legacy.get("学科")), class_name),
                "campus": s(legacy.get("校区")),
                "teacher": normalize_teacher(s(legacy.get("老师"))),
                "start_date": s(legacy.get("开课")),
                "end_date": s(legacy.get("结课")),
                "weekday": "",
                "time_range": s(legacy.get("时间")),
                "lecture_times": s(legacy.get("时间")),
                "capacity_text": s(legacy.get("已报预招")),
                "class_type": class_type(class_name),
                "raw": legacy,
                "last_import_batch": None,
                "active_in_latest": False,
            }
        if class_name:
            enrollments[legacy_eid] = {
                "eid": legacy_eid,
                "student_id": sid or None,
                "source_student_id": s(legacy.get("studentId") or legacy.get("sourceStudentId") or legacy.get("id")),
                "family_id": s(legacy.get("familyId")) or (students.get(sid, {}).get("family_id") if sid else None),
                "class_id": class_id,
                "student_name": s(legacy.get("姓名")),
                "phone": s(legacy.get("电话")),
                "class_name": class_name,
                "class_display_name": s(legacy.get("班级名称")) or class_name,
                "normalized_class_name": nclass,
                "grade": s(legacy.get("年级")),
                "subject": normalize_subject(s(legacy.get("学科")), class_name),
                "term": s(legacy.get("期")) or term_of(s(legacy.get("开课")), s(legacy.get("学期"))),
                "term_name": s(legacy.get("学期")),
                "campus": s(legacy.get("校区")),
                "teacher": normalize_teacher(s(legacy.get("老师"))),
                "time_range": s(legacy.get("时间")),
                "lecture_times": s(legacy.get("时间")),
                "start_date": s(legacy.get("开课")),
                "end_date": s(legacy.get("结课")),
                "source_status": s(legacy.get("源状态")) or s(legacy.get("状态")) or "未出现在本次导出",
                "assignment_status": s(legacy.get("分配状态")),
                "display_status": s(legacy.get("状态")),
                "capacity_text": s(legacy.get("已报预招")),
                "amount_due": money(legacy.get("应收") or legacy.get("课费")),
                "book_fee": money(legacy.get("书本费")),
                "arrears": money(legacy.get("欠费")),
                "fee_text": s(legacy.get("课费") or legacy.get("应收")),
                "is_void": bool(legacy.get("作废") or False),
                "is_manual": bool(legacy.get("手工") or False),
                "active_in_latest": False,
                "last_import_batch": None,
                "raw": legacy,
            }

    # Schedule: keep curated local schedule; tie to classes where names normalize.
    by_nclass = {item["normalized_class_name"]: item["id"] for item in classes.values()}
    schedule_items: Dict[str, Dict[str, Any]] = {}
    for idx, row in enumerate(legacy_schedule, start=1):
        class_name = s(row.get("班级名称") or row.get("课程") or row.get("班级"))
        nclass = normalized_class_name(class_name)
        schedule_id = s(row.get("班号")) or f"SCH-{stable_hash(idx, class_name, row.get('星期'), row.get('时间'))}"
        schedule_items[schedule_id] = {
            "schedule_id": schedule_id,
            "class_no": s(row.get("班号")),
            "class_id": by_nclass.get(nclass),
            "class_name": class_name,
            "normalized_class_name": nclass,
            "term": s(row.get("期")) or "2026秋",
            "weekday": s(row.get("星期")),
            "time_range": s(row.get("时间")),
            "course": s(row.get("课程")) or class_name,
            "subject": normalize_subject(s(row.get("学科")), class_name),
            "grade": s(row.get("年级")) or re.match(r"^([1-9])", class_name).group(1) + "年级" if re.match(r"^([1-9])", class_name) else s(row.get("年级")),
            "teacher": normalize_teacher(s(row.get("老师"))),
            "teacher_full_name": s(row.get("老师全名")) or normalize_teacher(s(row.get("老师"))),
            "campus": s(row.get("校区")),
            "room": s(row.get("教室")),
            "source": s(row.get("来源")) or "课表",
            "class_type": s(row.get("班型")) or class_type(class_name),
            "enrolled_count": int(row.get("在班人数") or row.get("人数") or 0),
            "start_date": s(row.get("开课")),
            "end_date": s(row.get("结课")),
            "raw": row,
            "last_import_batch": batch_id,
            "active_in_latest": True,
        }

    order_rows: Dict[str, Dict[str, Any]] = {}
    for idx, row in enumerate(legacy_orders, start=1):
        raw_id = s(row.get("单号"))
        oid = raw_id if raw_id and raw_id not in order_rows else f"ORD-{stable_hash(raw_id, row.get('商品'), row.get('姓名'), row.get('金额'), row.get('支付'), idx)}"
        order_rows[oid] = {
            "id": oid,
            "order_no": raw_id,
            "family_id": s(row.get("familyId")) or None,
            "child_id": s(row.get("childId")) or None,
            "source_student_id": s(row.get("sourceStudentId")),
            "student_name": s(row.get("姓名")),
            "phone": s(row.get("电话")),
            "campus": s(row.get("校区")),
            "teacher": s(row.get("老师")),
            "term": s(row.get("学期")),
            "product": s(row.get("商品")),
            "ordered_at": s(row.get("下单")),
            "paid_at": s(row.get("支付")),
            "amount": money(row.get("金额")),
            "payment_method": s(row.get("方式")),
            "payment_status": s(row.get("状态")),
            "assignment_status": s(row.get("分配状态")),
            "raw": row,
        }

    followups: List[Dict[str, Any]] = []
    for kind, block in [("expansion", legacy_state.get("expansion") or {}), ("renew", legacy_state.get("renewFollowup") or {})]:
        for key, value in block.items():
            parts = str(key).split("|", 1)
            term = parts[0] if parts else ""
            student_id = parts[1] if len(parts) > 1 else ""
            followups.append({
                "kind": kind,
                "term": term,
                "student_id": student_id or None,
                "status": s(value.get("状态")) if isinstance(value, dict) else s(value),
                "note": s(value.get("备注")) if isinstance(value, dict) else "",
                "next_followup_date": s(value.get("下次跟进")) if isinstance(value, dict) else "",
                "source_key": f"{kind}|{key}",
                "raw": value if isinstance(value, dict) else {"value": value},
            })

    leaves: List[Dict[str, Any]] = []
    for idx, item in enumerate(legacy_state.get("leaves") or [], start=1):
        lid = s(item.get("lid")) or f"L-{stable_hash(item.get('studentId'), item.get('班级'), item.get('日期'), idx)}"
        leaves.append({
            "lid": lid,
            "student_id": s(item.get("studentId")) or None,
            "student_name": s(item.get("姓名")),
            "class_name": s(item.get("班级")),
            "leave_date": s(item.get("日期")),
            "reason": s(item.get("原因")),
            "refund_amount": money(item.get("折算金额")),
            "note": s(item.get("备注")),
            "created_at_text": s(item.get("创建时间")),
            "raw": item,
        })

    class_progress = [
        {"class_name": s(k), "lecture_no": int(v or 0)}
        for k, v in (legacy_state.get("progress") or {}).items()
        if s(k)
    ]

    family_assignment_rules: List[Dict[str, Any]] = []
    for family_id, rules in (legacy_state.get("familyAssignmentRules") or {}).items():
        if not isinstance(rules, dict):
            continue
        for grade_key, child_id in rules.items():
            family_assignment_rules.append({"family_id": s(family_id), "grade_key": s(grade_key), "child_id": s(child_id)})

    op_logs: List[Dict[str, Any]] = []
    for idx, item in enumerate(legacy_state.get("opLog") or [], start=1):
        source_hash = stable_hash(item.get("时间"), item.get("动作"), item.get("对象"), item.get("变更"), idx, length=20)
        op_logs.append({
            "source_hash": source_hash,
            "logged_at": s(item.get("时间")),
            "action": s(item.get("动作")),
            "target": s(item.get("对象")),
            "class_name": s(item.get("班级")),
            "change": s(item.get("变更")),
            "detail": item,
        })

    latest_status = Counter(s(r.get("状态")) for r in rows)
    batch_summary.update({
        "roster_rows": len(rows),
        "unique_export_students": len({s(r.get("学生ID")) for r in rows if s(r.get("学生ID"))}),
        "unique_export_classes": len({s(r.get("班级名称")) for r in rows if s(r.get("班级名称"))}),
        "export_status_counts": dict(latest_status),
        "legacy_enrollments_preserved_for_review": sum(1 for e in enrollments.values() if not e.get("active_in_latest") and e.get("last_import_batch") is None),
        "schedule_rows": len(schedule_items),
        "orders_rows": len(order_rows),
        "followup_rows": len(followups),
        "leave_rows": len(leaves),
        "op_log_rows": len(op_logs),
    })

    return {
        "import_batches": [{
            "batch_id": batch_id,
            "source_name": ROSTER_FILE,
            "source_path": str(export_dir),
            "row_count": len(rows),
            "class_count": len({s(r.get("班级名称")) for r in rows if s(r.get("班级名称"))}),
            "student_count": len({s(r.get("学生ID")) for r in rows if s(r.get("学生ID"))}),
            "summary": batch_summary,
        }],
        "raw_roster_rows": raw_roster_rows,
        "families": list(families.values()),
        "students": list(students.values()),
        "classes": list(classes.values()),
        "enrollments": list(enrollments.values()),
        "orders": list(order_rows.values()),
        "schedule_items": list(schedule_items.values()),
        "course_outlines": [{"id": "main", "payload": outlines}],
        "followups": followups,
        "leaves": leaves,
        "class_progress": class_progress,
        "family_assignment_rules": family_assignment_rules,
        "op_logs": op_logs,
        "_summary": batch_summary,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--export-dir", default=str(DEFAULT_EXPORT_DIR))
    parser.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR))
    parser.add_argument("--out-dir", default=str(DEFAULT_ARCHIVE_DIR))
    args = parser.parse_args()

    export_dir = Path(args.export_dir)
    data_dir = Path(args.data_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    payloads = build_payloads(export_dir, data_dir)
    for key, value in payloads.items():
        if key.startswith("_"):
            continue
        (out_dir / f"{key}.json").write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "dry_run_report.json").write_text(json.dumps(payloads["_summary"], ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payloads["_summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
